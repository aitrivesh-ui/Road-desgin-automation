# -*- coding: utf-8 -*-
"""Build RoadAutomation_DataStarter.xlsx from csv/templates headers (requires openpyxl)."""
from __future__ import annotations

import csv
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install -r civil3d_automation/requirements-tools.txt", file=sys.stderr)
    sys.exit(1)

ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
TEMPL = os.path.join(ROOT, "csv", "templates")
OUT = os.path.join(TEMPL, "RoadAutomation_DataStarter.xlsx")

SHEETS = [
    ("README", None),
    ("alignment_pi", "alignment_pi.csv"),
    ("profile_pvis", "profile_pvis.csv"),
    ("section_widths", "section_widths.csv"),
    ("signage_schedule", "signage_schedule.csv"),
    ("payitems", "payitems.csv"),
]

HEADER_FILL = PatternFill("solid", fgColor="1c2030")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="F0A500")


def _read_headers(csv_name: str) -> list[str]:
    path = os.path.join(TEMPL, csv_name)
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return [h.strip() for h in next(csv.reader(f), [])]


def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, csv_name in SHEETS:
        ws = wb.create_sheet(title=sheet_name[:31])
        if csv_name is None:
            lines = [
                "Road automation — Excel data starter",
                "",
                "How to use:",
                "1. Fill data only on the sheets named after CSV files (do not rename sheets).",
                "2. When saving to CSV, use 'CSV UTF-8 (Comma delimited) (*.csv)' in Excel if available.",
                "3. Save each sheet as its matching file name under civil3d_automation/csv/ :",
                "   - alignment_pi.csv",
                "   - profile_pvis.csv",
                "   - section_widths.csv",
                "   - signage_schedule.csv",
                "   - payitems.csv",
                "4. Update paths in config/project.json if you use different file names.",
                "",
                "Regenerate this workbook after template column changes:",
                "  python tools/build_starter_workbook.py",
            ]
            for i, line in enumerate(lines, start=1):
                ws.cell(row=i, column=1, value=line)
            ws.column_dimensions["A"].width = 92
            continue

        headers = _read_headers(csv_name)
        ws.append(headers)
        for c, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.freeze_panes = "A2"
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
