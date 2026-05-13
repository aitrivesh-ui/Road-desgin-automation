# -*- coding: utf-8 -*-
"""
Export data sheets from RoadAutomation_DataStarter.xlsx (or compatible) to UTF-8 CSV files.

Usage:
  python export_workbook_to_csv.py path/to/workbook.xlsx path/to/output_csv_dir

Requires: pip install openpyxl (see requirements-tools.txt)
"""
from __future__ import annotations

import argparse
import csv
import os
import sys

# Sheet title -> output filename under csv/
SHEET_TO_CSV = {
    "alignment_pi": "alignment_pi.csv",
    "profile_pvis": "profile_pvis.csv",
    "section_widths": "section_widths.csv",
    "signage_schedule": "signage_schedule.csv",
    "payitems": "payitems.csv",
}


def export_workbook(workbook_path: str, out_dir: str) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("Install openpyxl: pip install -r civil3d_automation/requirements-tools.txt")

    log: list[str] = []
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    os.makedirs(out_dir, exist_ok=True)

    for sheet_name, csv_name in SHEET_TO_CSV.items():
        if sheet_name not in wb.sheetnames:
            log.append("WARN: Sheet not found, skipped: " + sheet_name)
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        out_path = os.path.join(out_dir, csv_name)
        with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            for row in rows_iter:
                if row is None:
                    continue
                cells = list(row)
                if not any(c is not None and str(c).strip() != "" for c in cells):
                    continue
                w.writerow([("" if c is None else c) for c in cells])
        log.append("OK: Wrote " + out_path)

    wb.close()
    return log


def main() -> int:
    ap = argparse.ArgumentParser(description="Export starter workbook sheets to CSV files.")
    ap.add_argument("workbook", help="Path to .xlsx")
    ap.add_argument("out_dir", help="Folder for CSV outputs (e.g. civil3d_automation/csv)")
    args = ap.parse_args()
    if not os.path.isfile(args.workbook):
        print("ERROR: Workbook not found: " + args.workbook, file=sys.stderr)
        return 1
    try:
        for line in export_workbook(args.workbook, args.out_dir):
            print(line)
        return 0
    except Exception as ex:
        print("ERROR: %s" % ex, file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
